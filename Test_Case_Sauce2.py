from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait 
from selenium.webdriver.support import expected_conditions as ec 
import pytest
import openpyxl
from constants import globalConstants as c


class Test_Case_Homework2:
    def setup_method(self): 
        self.driver = webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window() 

    def teardown_method(self):
        self.driver.quit()
    

    def getData1():
        excel = openpyxl.load_workbook(c.invalid_login_xlsx)
        sheet = excel["Sheet1"] #hangi sayfada çalışacağımı gösteriyorum
        rows = sheet.max_row #kaçıncı satıra kadar veri var?
        data = []
        for i in range(2,rows+1):
            username = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((username,password))

        return data
    
    def getData2():
        excel = openpyxl.load_workbook(c.valid_login_xlsx)
        sheet = excel["Sheet1"] #hangi sayfada çalışacağımı gösteriyorum
        rows = sheet.max_row #kaçıncı satıra kadar veri var?
        data = []
        for i in range(2,rows+1):
            username = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((username,password))

        return data


    @pytest.mark.parametrize("username,password",getData1())
    def test_invalid_login(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys(username)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys(password)
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        errorMessage = self.driver.find_element(By.XPATH,c.ERROR_MESSAGE_XPATH)
        assert errorMessage.text == c.USERNAME_PASSWORD_DONT_MATCH

        
    @pytest.mark.parametrize("username,password",getData2())
    def test_valid_remove(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys(username)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys(password)
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        self.driver.execute_script("window.scrollTo(0,500)") 
        addToCart = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.ADD_TO_CHART_XPATH)))
        addToCart.click()
        cartControl = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.CART_CONTROL_XPATH)))
        cartControl.click()
        remove = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.REMOVE_ITEM_XPATH)))
        assert remove.text == c.REMOVE_ITEM
 
    

    @pytest.mark.parametrize("username,password",getData2())
    def test_valid_purchase(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys(username)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys(password)
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        self.driver.execute_script("window.scrollTo(0,500)") 
        addToCart = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.ADD_TO_CHART_XPATH)))
        addToCart.click()
        cartControl = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.CART_CONTROL_XPATH)))
        cartControl.click()
        successCheckout = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.SUCCESS_CHECKOUT_XPATH)))
        successCheckout.click()
        firstnameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.FIRSTNAME_ID)))
        firstnameInput.send_keys("Zehra")
        lastnameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.LASTNAME_ID)))
        lastnameInput.send_keys("Temizel")
        postalcodeInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.POSTAL_CODE)))
        postalcodeInput.send_keys("3344")  
        successContinue = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.SUCCESS_CONTINUE_XPATH)))
        successContinue.click()
        successFinish = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.SUCCESS_FINISH_XPATH)))
        successFinish.click()
        finishMessage = self.driver.find_element(By.XPATH,c.FINISH_MESSAGE_XPATH)
        assert finishMessage.text == c.FINISH_MESSAGE




    @pytest.mark.parametrize("username,password",getData2())
    def test_valid_checkout(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.USERNAME_ID)))
        usernameInput.send_keys(username)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.PASSWORD_ID)))
        passwordInput.send_keys(password)
        loginButton = self.driver.find_element(By.ID,c.LOGIN_BUTTON_ID)
        loginButton.click()
        self.driver.execute_script("window.scrollTo(0,500)") 
        addToCart = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.ADD_TO_CHART_XPATH)))
        addToCart.click()
        cartControl = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.CART_CONTROL_XPATH)))
        cartControl.click()
        successCheckout = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.SUCCESS_CHECKOUT_XPATH)))
        successCheckout.click()
        finishMessage1 = self.driver.find_element(By.XPATH,c.FINISH_MESSAGE_XPATH2)
        assert finishMessage1.text == c.CHECKOUT_MESSAGE
 
    
